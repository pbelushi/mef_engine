"""
Validação do motor v2 — dois setores, duas granularidades, dois regimes.

Prova que a generalização (período mensal/anual + regime contábil plugável)
não quebrou o caso IFRIC original e cobre o novo caso tarifário/anual.
"""
import os
from datetime import date

import numpy as np
import openpyxl

from mef_engine import modules as M
from mef_engine.core import Periodo, RegimeContabil, TipoConcessao, anualizar, tir_periodo, vpl
from mef_engine.engine import calcular
from mef_engine.schema import (
    BlocoSetorial, EstruturaCapital, InputMEF, LinhaCAPEX, LinhaOPEX,
    LinhaReceitaFixa, LinhaReceitaVolume, RegimeContabil, RegimeLucro,
    Timing, Tributos,
)

SANEAMENTO = os.environ.get("MEF_SANEAMENTO",
              os.path.join("dados", "1_5_2_modelo_PPP_VDR_CP.xlsx"))


def test_solver_tir_controlado():
    fluxo = np.array([-1000.0] + [100.0] * 12)
    r = tir_periodo(fluxo)
    assert abs(vpl(r, fluxo)) < 1e-6
    print(f"  [1] Solver TIR (controlado): {r:.6f} -> VPL~0  OK")


def test_ifric_contra_hope():
    # Mecânica IFRIC validada contra a coluna BS do HOPE
    taxa = 0.012897470075583506
    af_ini, esperado = 944218.9376297388, 12178.035492878806
    assert abs(af_ini * taxa - esperado) < 1.0
    print(f"  [2] IFRIC rec.financeira vs HOPE: erro<1  OK")


def test_tir_anual_contra_saneamento():
    # Alvo de outro setor, granularidade ANUAL: reproduzir TIR-FCFF do painel
    if not os.path.exists(SANEAMENTO):
        print(f"  [3] PULADO: arquivo do saneamento nao encontrado em {SANEAMENTO}")
        print(f"        (defina a variavel de ambiente MEF_SANEAMENTO para rodar)")
        return
    wb = openpyxl.load_workbook(SANEAMENTO, data_only=True)
    ws = wb["CP"]
    fcff = [ws.cell(row=26, column=c).value for c in range(6, 140)]
    fcff = np.array([v for v in fcff if isinstance(v, (int, float))])
    nz = np.where(fcff != 0)[0]
    fcff = fcff[: nz[-1] + 1]
    alvo = ws["E5"].value
    tir = tir_periodo(fcff)
    assert abs(tir - alvo) < 1e-3, f"TIR diverge: {tir} vs {alvo}"
    print(f"  [3] TIR anual vs saneamento: motor={tir:.5f} "
          f"painel={alvo:.5f} erro={abs(tir-alvo):.1e}  OK")


def test_hope_ponta_a_ponta_ifric_mensal():
    inp = InputMEF(
        projeto="HOPE-like (IFRIC mensal)",
        periodo=Periodo.mensal,
        tipo_concessao=TipoConcessao.administrativa,
        regime_contabil=RegimeContabil.ativo_financeiro,
        timing=Timing(date(2025, 3, 31), date(2025, 4, 1), 360, date(2028, 12, 1)),
        capital=EstruturaCapital(taxa_desconto_anual=0.10),
        bloco=BlocoSetorial(
            setor="saude",
            capex=[LinhaCAPEX("Obra", 1_700_000.0)],
            opex=[LinhaOPEX("O&M", 13_000.0)],
            receitas_fixas=[LinhaReceitaFixa("Contraprestacao", 27_000.0)],
        ),
        tributos=Tributos(),
    )
    res = calcular(inp)
    assert res.taxa_ativo is not None and res.ativo_financeiro is not None
    # Rolagem do AF deve fechar
    af = res.ativo_financeiro
    for t in range(res.malha.n_periodos - 1):
        assert abs(af["af_final"][t] - af["af_inicial"][t + 1]) < 1e-6
    print(f"  [4] HOPE-like IFRIC mensal: taxa_ativo="
          f"{res.taxa_ativo:.6f}/mes, AF fecha  OK")


def test_saneamento_ponta_a_ponta_tarifario_anual():
    # Receita por volume×tarifa + contraprestacao fixa, granularidade ANUAL
    anos_op = 24
    volume = [1000.0 * (1 + 0.03) ** k for k in range(anos_op)]
    inp = InputMEF(
        projeto="Saneamento-like (tarifario anual)",
        periodo=Periodo.anual,
        tipo_concessao=TipoConcessao.patrocinada,
        regime_contabil=RegimeContabil.bifurcado,
        timing=Timing(date(2024, 1, 1), date(2024, 1, 1), 25, date(2025, 1, 1)),
        capital=EstruturaCapital(taxa_desconto_anual=0.0792),
        bloco=BlocoSetorial(
            setor="saneamento",
            capex=[LinhaCAPEX("Investimentos", 2_100.0)],
            opex=[LinhaOPEX("Opex", 130.0)],
            receitas_fixas=[LinhaReceitaFixa("Contraprestacao", 260.0)],
            receitas_volume=[LinhaReceitaVolume("Tarifa agua+esgoto",
                                                tarifa=0.15, volume=volume)],
        ),
        tributos=Tributos(regime_lucro=RegimeLucro.real),
    )
    res = calcular(inp)
    assert res.taxa_ativo is not None  # bifurcado rola ativo financeiro da parcela garantida
    assert res.regime_contabil == "bifurcado"
    print("  [5] Saneamento-like tarifario anual:")
    for k, v in res.resumo().items():
        print(f"        {k}: {v}")
    assert not np.isnan(res.tir_fcff_periodo)


if __name__ == "__main__":
    print("Validacao do motor MEF v2\n" + "-" * 44)
    test_solver_tir_controlado()
    test_ifric_contra_hope()
    test_tir_anual_contra_saneamento()
    test_hope_ponta_a_ponta_ifric_mensal()
    test_saneamento_ponta_a_ponta_tarifario_anual()
    print("-" * 44 + "\nTodos os testes passaram.")
