"""
Ingestão de CAPEX/OPEX a partir de planilhas VARIADAS.

Estratégia (decisão do projeto): não impor formato rígido. O parser detecta a
estrutura por heurística e ANCORA na linha de soma/total — a parte mais estável
e identificável de qualquer MEF. Em seguida RECONCILIA a soma dos itens contra
o total declarado: se baterem (dentro de tolerância), a leitura é confiável; se
não, sinaliza para revisão humana em vez de ingerir dado silenciosamente errado.

Sem dependência de formato de coluna: descobre a coluna de rótulos (texto) e a
de valores (numérica) por varredura, e a linha de total por palavra-chave.
"""
from __future__ import annotations

import csv
import io
import re
import unicodedata
from dataclasses import dataclass, field

import openpyxl

PALAVRAS_TOTAL = ("total", "soma", "subtotal", "totais")
PALAVRAS_RUIDO = ("data-base", "tir", "vpl", "taxa", "prazo", "moeda")
PADRAO_PERIODO_ROTULO = re.compile(r"(?:ano|m[eê]s|per[ií]odo)\s*0*(\d+)", re.IGNORECASE)


def _norm(s) -> str:
    if not isinstance(s, str):
        return ""
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return s.strip().lower()


class _Celula:
    __slots__ = ("value",)

    def __init__(self, value):
        self.value = value


class _GradeLista:
    """Adaptador que expõe a mesma interface mínima de uma worksheet openpyxl
    (`.cell(row=, column=).value`, `.max_row`, `.max_column`) sobre uma lista
    de listas em memória — permite reusar as heurísticas de detecção de
    seção (pensadas para openpyxl) também em arquivos .csv/.xls, que não têm
    um objeto worksheet nativo equivalente."""

    def __init__(self, linhas: list[list]):
        self._linhas = linhas
        self.max_row = len(linhas)
        self.max_column = max((len(l) for l in linhas), default=0)

    def cell(self, row, column):
        linha = self._linhas[row - 1] if 0 < row <= self.max_row else []
        valor = linha[column - 1] if 0 < column <= len(linha) else None
        return _Celula(valor)


_PADRAO_NUM_BR = re.compile(r"^-?\d{1,3}(\.\d{3})*,\d+$|^-?\d+,\d+$")


def _paresear_numero_csv(s: str):
    """Converte uma célula textual de .csv em float quando plausível — aceita
    tanto o formato BR (1.234,56) quanto o internacional (1234.56); devolve a
    string original (rótulo) se não for reconhecível como número."""
    s = s.strip()
    if not s:
        return None
    if _PADRAO_NUM_BR.match(s):
        return float(s.replace(".", "").replace(",", "."))
    try:
        return float(s)
    except ValueError:
        return s


def _ler_csv_generico(conteudo: bytes) -> list[list]:
    """Decodifica bytes de .csv (utf-8 com BOM ou, na falha, latin-1 — comum
    em export do Excel BR) e detecta o delimitador (',' ou ';') por
    amostragem, em vez de impor um separador fixo."""
    for enc in ("utf-8-sig", "latin-1"):
        try:
            texto = conteudo.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        texto = conteudo.decode("utf-8", errors="replace")
    amostra = texto[:2048]
    try:
        delimitador = csv.Sniffer().sniff(amostra, delimiters=",;\t").delimiter
    except csv.Error:
        delimitador = ";" if amostra.count(";") > amostra.count(",") else ","
    leitor = csv.reader(io.StringIO(texto), delimiter=delimitador)
    return [[_paresear_numero_csv(c) if isinstance(c, str) else c for c in linha]
            for linha in leitor]


def _ler_xls_generico(conteudo: bytes) -> list[list]:
    import xlrd
    livro = xlrd.open_workbook(file_contents=conteudo)
    planilha = livro.sheet_by_index(0)
    return [[planilha.cell_value(r, c) or None for c in range(planilha.ncols)]
            for r in range(planilha.nrows)]


def carregar_grade(arquivo, nome_arquivo: str, aba: str | None = None):
    """Carrega qualquer um dos formatos aceitos (.xlsx/.xlsm via openpyxl,
    .xls via xlrd, .csv via csv) numa interface única de leitura por célula —
    o resto do módulo (detecção de seção, reconciliação) não precisa saber
    qual formato originou os dados.

    `arquivo` é um caminho (str) ou um objeto com `.read()` (ex.: upload do
    Streamlit); `nome_arquivo` só serve para decidir o formato pela extensão
    quando `arquivo` não é, ele mesmo, o caminho.
    """
    ext = nome_arquivo.rsplit(".", 1)[-1].lower()
    if ext in ("xlsx", "xlsm"):
        wb = openpyxl.load_workbook(arquivo, data_only=True)
        return wb[aba] if aba else wb[wb.sheetnames[0]]
    conteudo = arquivo.read() if hasattr(arquivo, "read") else open(arquivo, "rb").read()
    if ext == "xls":
        return _GradeLista(_ler_xls_generico(conteudo))
    if ext == "csv":
        return _GradeLista(_ler_csv_generico(conteudo))
    raise ValueError(f"Formato não suportado: .{ext} (use .xlsx, .xls ou .csv)")


@dataclass
class LinhaItem:
    nome: str
    valor: float
    linha: int
    curva: dict = field(default_factory=dict)  # período relativo (0-based) -> valor


@dataclass
class SecaoIngerida:
    titulo: str
    itens: list = field(default_factory=list)
    total_declarado: float | None = None
    linha_total: int | None = None

    @property
    def soma_itens(self) -> float:
        return sum(i.valor for i in self.itens)

    def reconciliar(self, tol_rel: float = 1e-3) -> dict:
        """Compara soma dos itens com o total declarado. A checagem central de
        qualidade: dá confiança de que o parser leu o bloco certo."""
        if self.total_declarado is None:
            return {"ok": None, "motivo": "sem linha de total para ancorar"}
        soma = self.soma_itens
        if self.total_declarado == 0:
            ok = abs(soma) < 1.0
            return {"ok": ok, "soma": soma, "total": self.total_declarado,
                    "erro_rel": None}
        erro_rel = abs(soma - self.total_declarado) / abs(self.total_declarado)
        return {"ok": erro_rel <= tol_rel, "soma": soma,
                "total": self.total_declarado, "erro_rel": erro_rel}


def _primeiro_rotulo(ws, r, max_col):
    """1ª célula de texto (>2 chars) da linha = rótulo candidato, com sua
    coluna. (None, None) se a linha não tem nenhuma célula de texto assim."""
    for c in range(1, max_col + 1):
        v = ws.cell(row=r, column=c).value
        if isinstance(v, str) and len(v.strip()) > 2:
            return v.strip(), c
    return None, None


def _melhor_valor_na_linha(ws, r, c_inicio, c_fim):
    """Maior valor numérico plausível à direita do rótulo (ignora colunas que
    repetem o mesmo número em unidades diferentes — pega o primeiro numérico)."""
    for c in range(c_inicio, c_fim + 1):
        v = ws.cell(row=r, column=c).value
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            return float(v), c
    return None, None


def _periodo_da_celula(v):
    """Interpreta uma célula de cabeçalho como índice de período: 'Ano 1'
    -> 0, 'Mês 3' -> 2, ano absoluto (1900-2100) -> o próprio ano (ainda não
    normalizado — `detectar_cabecalho_periodos` subtrai o mínimo do grupo).
    None se não bater em nenhum padrão reconhecido — não arrisca confundir
    texto/número solto com período."""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        ano = int(v)
        if ano == v and 1900 <= ano <= 2100:
            return ano
        return None
    if isinstance(v, str):
        m = PADRAO_PERIODO_ROTULO.search(v)
        if m:
            return int(m.group(1)) - 1
    return None


def detectar_cabecalho_periodos(ws, r, c_inicio, c_fim):
    """Cabeçalho de período na linha `r` (ex.: 'Ano 1'|'Ano 2'|'Ano 3', ou
    2024|2025|2026): exige >=2 colunas com período reconhecido, em sequência
    ESTRITAMENTE crescente e com passo constante — filtro deliberadamente
    conservador contra falso positivo (números ou textos soltos não formam
    cabeçalho). Retorna {coluna: índice_relativo_0_based} ou None; usado só
    para decidir se uma linha de item tem CURVA de desembolso (vários
    valores = períodos) em vez da mesma cifra em outra unidade."""
    achados = [(c, p) for c in range(c_inicio, c_fim + 1)
              for p in [_periodo_da_celula(ws.cell(row=r, column=c).value)]
              if p is not None]
    if len(achados) < 2:
        return None
    valores = [p for _, p in achados]
    if valores != sorted(set(valores)) or len(set(valores)) != len(valores):
        return None
    passos = {valores[i + 1] - valores[i] for i in range(len(valores) - 1)}
    if len(passos) != 1:
        return None
    base = valores[0]
    return {c: p - base for c, p in achados}


def _curva_da_linha(ws, r, cab_periodos):
    curva = {}
    for c, k in cab_periodos.items():
        v = ws.cell(row=r, column=c).value
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            curva[k] = float(v)
    return curva


def _candidata_cabecalho(ws, r, max_col):
    """Rótulo da linha SE ela for candidata a cabeçalho de seção: tem texto,
    não é ruído/total, e NÃO tem valor numérico na linha — é o que distingue
    um título de seção de uma linha de item (rótulo + valor). None se não for
    candidata."""
    rotulo, c_rotulo = _primeiro_rotulo(ws, r, max_col)
    if rotulo is None:
        return None
    nr = _norm(rotulo)
    if any(p in nr for p in PALAVRAS_RUIDO) or any(p in nr for p in PALAVRAS_TOTAL):
        return None
    valor, _ = _melhor_valor_na_linha(ws, r, c_rotulo + 1, max_col)
    return rotulo if valor is None else None


def ingerir_secao(ws, linha_ini: int, linha_fim: int,
                  titulo: str = "") -> SecaoIngerida:
    """Lê itens numa faixa de linhas e detecta a linha de total por
    palavra-chave. Se a linha imediatamente anterior à faixa for um
    cabeçalho de período reconhecido (`detectar_cabecalho_periodos`), cada
    item ganha uma `curva` (período relativo -> valor) lida nessas colunas,
    e `valor` passa a ser a soma da curva — uma curva de desembolso real,
    não a mesma cifra repetida em outra unidade (que é o caso sem cabeçalho,
    onde só o 1º valor numérico é lido, como sempre)."""
    sec = SecaoIngerida(titulo=titulo)
    max_col = min(ws.max_column, 30)
    cab_periodos = (detectar_cabecalho_periodos(ws, linha_ini - 1, 1, max_col)
                    if linha_ini > 1 else None)
    for r in range(linha_ini, linha_fim + 1):
        rotulo, c_rotulo = _primeiro_rotulo(ws, r, max_col)
        if rotulo is None:
            continue
        nr = _norm(rotulo)
        if any(p in nr for p in PALAVRAS_RUIDO):
            continue
        if any(p in nr for p in PALAVRAS_TOTAL):
            valor, _ = _melhor_valor_na_linha(ws, r, c_rotulo + 1, max_col)
            if valor is not None:
                sec.total_declarado = valor
                sec.linha_total = r
            continue
        curva = _curva_da_linha(ws, r, cab_periodos) if cab_periodos else {}
        if curva:
            valor = sum(curva.values())
        else:
            valor, _ = _melhor_valor_na_linha(ws, r, c_rotulo + 1, max_col)
        if valor is None:
            continue
        sec.itens.append(LinhaItem(nome=rotulo, valor=valor, linha=r, curva=curva))
    return sec


def detectar_faixas(ws, linha_ini: int = 1, linha_fim: int | None = None,
                    max_col: int = 30) -> list[tuple[int, int, str]]:
    """Detecta automaticamente as faixas (ini, fim, título) de seções numa
    aba, por heurística de CABEÇALHO: candidata é uma linha com rótulo de
    texto e SEM valor numérico (uma linha de item sempre tem rótulo+valor,
    um cabeçalho não). A faixa de cada candidata vai até a próxima candidata
    (ou o fim da varredura).

    Uma candidata só é aceita como seção REAL se a faixa resultante tiver ao
    menos 1 item E uma linha de total para ancorar — mesma filosofia do
    projeto (ancorar no total antes de confiar na leitura). Candidatas sem
    isso são ruído (nota, título de página, parâmetro solto) e são
    descartadas silenciosamente; a ausência na lista resultante é o sinal."""
    max_col = min(ws.max_column, max_col)
    linha_fim = linha_fim or ws.max_row
    candidatos = []
    for r in range(linha_ini, linha_fim + 1):
        titulo = _candidata_cabecalho(ws, r, max_col)
        if titulo is not None:
            candidatos.append((r, titulo))

    faixas = []
    for i, (r_header, titulo) in enumerate(candidatos):
        proxima = candidatos[i + 1][0] if i + 1 < len(candidatos) else linha_fim + 1
        sec = ingerir_secao(ws, r_header + 1, proxima - 1, titulo)
        if sec.itens and sec.total_declarado is not None:
            faixas.append((r_header + 1, sec.linha_total, titulo))
    return faixas


def detectar_e_ingerir(caminho: str, aba: str | None = None,
                       faixas: list[tuple[int, int, str]] | None = None) -> list:
    """Ingere múltiplas seções de uma aba. `faixas` = [(ini, fim, titulo), ...]
    explícitas, ou None (default) para detectar automaticamente por
    cabeçalho via `detectar_faixas` — útil quando a heurística não bate ou
    para focar um teste no parser+reconciliação isoladamente."""
    ws = carregar_grade(caminho, caminho, aba)
    if faixas is None:
        faixas = detectar_faixas(ws)
    return [ingerir_secao(ws, ini, fim, tit) for ini, fim, tit in faixas]


def ingerir_arquivo_secao_unica(arquivo, nome_arquivo: str, titulo: str = "") -> SecaoIngerida:
    """Ingestão para upload dedicado a um único bloco (ex.: CAPEX e OPEX
    enviados como arquivos separados na interface web) — não precisa
    detectar limites de seção por cabeçalho como `detectar_faixas`, porque o
    arquivo inteiro JÁ é a seção; só lê a grade inteira como um bloco de
    itens + (opcionalmente) uma linha de total para reconciliação."""
    ws = carregar_grade(arquivo, nome_arquivo)
    return ingerir_secao(ws, 1, ws.max_row, titulo=titulo)
