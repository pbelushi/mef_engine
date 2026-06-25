"""
Parsing assistido: fallback de IA quando a heurística de ingestão
(`ingest.planilha.detectar_faixas`) não acha NENHUMA seção. A IA só SUGERE
faixas (início, fim, título); a sugestão passa pela MESMA `ingerir_secao` —
e portanto pela mesma reconciliação soma-vs-total — que qualquer faixa
manual ou heurística. Nunca ingere dado sem essa checagem, com ou sem IA.
"""
from __future__ import annotations

import openpyxl

from ..ingest.planilha import detectar_faixas, ingerir_secao
from .cliente import IAIndisponivel
from .cliente import gerar_texto as _gerar_texto_padrao


def _dump_linhas(ws, linha_ini: int, linha_fim: int, max_col: int) -> str:
    linhas = []
    for r in range(linha_ini, linha_fim + 1):
        celulas = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        celulas = [c for c in celulas if c is not None]
        if celulas:
            linhas.append(f"{r}: {celulas}")
    return "\n".join(linhas)


def montar_prompt_parsing(ws, linha_ini: int = 1, linha_fim: int | None = None,
                          max_col: int = 10) -> str:
    """Prompt determinístico a partir do dump de linhas — testável sem rede."""
    linha_fim = linha_fim or ws.max_row
    dump = _dump_linhas(ws, linha_ini, linha_fim, max_col)
    return (
        "Esta é uma planilha de modelo econômico-financeiro (CAPEX/OPEX de "
        "uma concessão). Cada linha abaixo está no formato "
        "'NÚMERO_DA_LINHA: [valores das células, na ordem das colunas]'. "
        "Identifique as seções de itens (cada seção tem um título, uma "
        "faixa de linhas de itens, e termina numa linha de total/soma). "
        "Responda SOMENTE com uma linha por seção, no formato exato "
        "'INICIO|FIM|TÍTULO' (início e fim incluem a linha de total), sem "
        "nenhum texto antes ou depois.\n\n" + dump
    )


def parsear_resposta_parsing(texto: str) -> list[tuple[int, int, str]]:
    """Converte a resposta da IA ('INICIO|FIM|TÍTULO' por linha) no mesmo
    formato de `detectar_faixas`. Linhas mal-formadas são ignoradas — uma
    resposta ruim produz menos faixas, nunca derruba a ingestão (e as
    faixas que sobram ainda passam pela reconciliação)."""
    faixas = []
    for linha in texto.strip().splitlines():
        partes = [p.strip() for p in linha.strip().split("|")]
        if len(partes) != 3:
            continue
        try:
            ini, fim = int(partes[0]), int(partes[1])
        except ValueError:
            continue
        if partes[2]:
            faixas.append((ini, fim, partes[2]))
    return faixas


def sugerir_faixas(ws, linha_ini: int = 1, linha_fim: int | None = None,
                   max_col: int = 10, gerar_texto=_gerar_texto_padrao
                   ) -> list[tuple[int, int, str]]:
    """Sugestão de faixas via IA. Devolve [] (nunca levanta erro) se a IA
    estiver indisponível — quem chama trata como 'nada detectado'."""
    prompt = montar_prompt_parsing(ws, linha_ini, linha_fim, max_col)
    try:
        texto = gerar_texto(prompt)
    except IAIndisponivel:
        return []
    return parsear_resposta_parsing(texto)


def detectar_e_ingerir_com_ia_fallback(caminho: str, aba: str,
                                       linha_ini: int = 1,
                                       linha_fim: int | None = None,
                                       max_col: int = 10,
                                       gerar_texto=_gerar_texto_padrao) -> list:
    """Mesmo retorno de `ingest.planilha.detectar_e_ingerir`, mas com
    fallback de IA: roda a heurística primeiro; só chama a IA se ela não
    achar NENHUMA faixa. A sugestão da IA é ingerida pela mesma
    `ingerir_secao` — sujeita à mesma reconciliação que qualquer faixa."""
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb[aba]
    faixas = detectar_faixas(ws, linha_ini, linha_fim, max_col=max(max_col, 30))
    if not faixas:
        faixas = sugerir_faixas(ws, linha_ini, linha_fim, max_col, gerar_texto)
    return [ingerir_secao(ws, ini, fim, tit) for ini, fim, tit in faixas]
