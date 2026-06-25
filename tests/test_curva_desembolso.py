"""
Validação da ingestão de CAPEX/OPEX com curvas de desembolso (v3.7).

Decisão de design (deliberadamente conservadora): uma linha de item com
múltiplos valores numéricos só é tratada como CURVA (valores por período) se
houver um cabeçalho de período reconhecido (ex.: "Ano 1"|"Ano 2"|"Ano 3", ou
anos absolutos) imediatamente acima da faixa. Sem esse cabeçalho, múltiplos
números na mesma linha continuam sendo lidos como antes — só o 1º valor,
porque podem ser a mesma cifra em unidades diferentes (R$ mil vs. R$), não
uma curva. Errar essa distinção seria ingerir dado financeiro silenciosamente
errado — o que o módulo de ingestão foi desenhado para evitar.
"""
import os
import tempfile

import numpy as np
import openpyxl

from mef_engine.core import MalhaTemporal, Periodo
from mef_engine.ingest import secao_para_capex, secao_para_opex
from mef_engine.ingest.planilha import (
    detectar_cabecalho_periodos, ingerir_secao,
)
from mef_engine.modules import vetor_capex, vetor_opex
from mef_engine.schema import BlocoSetorial, InputMEF, LinhaOPEX


def test_detectar_cabecalho_periodos_rotulos_e_anos_absolutos():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Item", "Ano 1", "Ano 2", "Ano 3"])
    ws.append(["Item", 2024, 2025, 2026])
    ws.append(["Item", "nota", "outra coisa"])       # sem período -> None
    ws.append(["Item", 100, 50])                      # números soltos (não-período) -> None

    cab1 = detectar_cabecalho_periodos(ws, 1, 1, 4)
    cab2 = detectar_cabecalho_periodos(ws, 2, 1, 4)
    cab3 = detectar_cabecalho_periodos(ws, 3, 1, 4)
    cab4 = detectar_cabecalho_periodos(ws, 4, 1, 4)
    assert cab1 == {2: 0, 3: 1, 4: 2}
    assert cab2 == {2: 0, 3: 1, 4: 2}
    assert cab3 is None
    assert cab4 is None
    print("  [1] Detecta 'Ano N' e anos absolutos em sequência; rejeita o resto  OK")


def test_item_sem_cabecalho_mantem_comportamento_anterior():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["CAPEX X"])
    ws.append(["Obra", 1000.0, 1000000.0])   # mesma cifra em R$ mil e R$ — SEM cabeçalho
    ws.append(["Total CAPEX X", 1000.0])
    sec = ingerir_secao(ws, 2, 3, "CAPEX X")
    assert len(sec.itens) == 1
    assert sec.itens[0].valor == 1000.0   # só o 1º valor, como antes
    assert sec.itens[0].curva == {}
    print("  [2] Sem cabeçalho de período, múltiplos números não viram curva (1º valor, como antes)  OK")


def test_item_com_cabecalho_de_periodo_vira_curva():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["CAPEX Y"])
    ws.append(["Item", "Ano 1", "Ano 2", "Ano 3"])
    ws.append(["Obra", 100.0, 200.0, 50.0])
    ws.append(["Total CAPEX Y", 350.0])
    sec = ingerir_secao(ws, 3, 4, "CAPEX Y")
    assert len(sec.itens) == 1
    item = sec.itens[0]
    assert item.curva == {0: 100.0, 1: 200.0, 2: 50.0}
    assert item.valor == 350.0
    rec = sec.reconciliar()
    assert rec["ok"]
    print(f"  [3] Com cabeçalho de período, curva = {item.curva}, valor = soma = {item.valor}  OK")


def test_secao_para_capex_converte_curva_em_fracoes():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["CAPEX Y"])
    ws.append(["Item", "Ano 1", "Ano 2"])
    ws.append(["Obra", 300.0, 700.0])
    ws.append(["Total CAPEX Y", 1000.0])
    sec = ingerir_secao(ws, 3, 4, "CAPEX Y")
    linhas = secao_para_capex(sec)
    assert len(linhas) == 1
    assert linhas[0].valor_total == 1000.0
    assert linhas[0].curva == {0: 0.3, 1: 0.7}
    print(f"  [4] secao_para_capex converte curva absoluta em fração: {linhas[0].curva}  OK")


def test_secao_para_opex_usa_curva_absoluta_sem_extrapolar():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["OPEX Z"])
    ws.append(["Item", "Ano 1", "Ano 2"])
    ws.append(["O&M", 10.0, 20.0])
    ws.append(["Total OPEX Z", 30.0])
    sec = ingerir_secao(ws, 3, 4, "OPEX Z")
    linhas = secao_para_opex(sec)
    assert linhas[0].valor_periodo == 0.0
    assert linhas[0].curva == {0: 10.0, 1: 20.0}
    print(f"  [5] secao_para_opex: valor_periodo=0.0 (sem extrapolar), curva={linhas[0].curva}  OK")


def test_vetor_opex_curva_sobrepoe_e_curva_vazia_preserva_comportamento():
    # Caso 1: curva vazia (default) reproduz o comportamento anterior (constante).
    linha_constante = LinhaOPEX(nome="Opex", valor_periodo=10.0, periodo_inicio=0)
    bloco1 = BlocoSetorial(opex=[linha_constante])
    from mef_engine.core import TipoConcessao, RegimeContabil
    from mef_engine.schema import EstruturaCapital, Timing, Tributos
    from datetime import date
    timing = Timing(date(2024, 1, 1), date(2024, 1, 1), 4, date(2024, 1, 1))
    inp1 = InputMEF(projeto="x", periodo=Periodo.anual, tipo_concessao=TipoConcessao.comum,
                    regime_contabil=RegimeContabil.intangivel, timing=timing,
                    capital=EstruturaCapital(), bloco=bloco1, tributos=Tributos())
    malha = MalhaTemporal(inicio=timing.inicio_ppp, n_periodos=4, periodo=Periodo.anual)
    v1 = vetor_opex(inp1, malha)
    assert np.array_equal(v1, [10.0, 10.0, 10.0, 10.0])

    # Caso 2: curva sobrepõe períodos específicos, resto continua com valor_periodo.
    linha_curva = LinhaOPEX(nome="Opex", valor_periodo=10.0, periodo_inicio=0,
                            curva={1: 99.0})
    bloco2 = BlocoSetorial(opex=[linha_curva])
    inp2 = InputMEF(projeto="y", periodo=Periodo.anual, tipo_concessao=TipoConcessao.comum,
                    regime_contabil=RegimeContabil.intangivel, timing=timing,
                    capital=EstruturaCapital(), bloco=bloco2, tributos=Tributos())
    v2 = vetor_opex(inp2, malha)
    assert np.array_equal(v2, [10.0, 99.0, 10.0, 10.0])
    print(f"  [6] vetor_opex: curva vazia preserva o comportamento anterior {v1}; "
          f"curva sobrepõe período específico {v2}  OK")


if __name__ == "__main__":
    print("Validação da ingestão de CAPEX/OPEX com curvas de desembolso\n" + "-" * 48)
    test_detectar_cabecalho_periodos_rotulos_e_anos_absolutos()
    test_item_sem_cabecalho_mantem_comportamento_anterior()
    test_item_com_cabecalho_de_periodo_vira_curva()
    test_secao_para_capex_converte_curva_em_fracoes()
    test_secao_para_opex_usa_curva_absoluta_sem_extrapolar()
    test_vetor_opex_curva_sobrepoe_e_curva_vazia_preserva_comportamento()
    print("-" * 48 + "\nTodos os testes passaram.")
