"""
Validação da detecção automática de faixas de seção por cabeçalho (v3.6).

Antes, `detectar_e_ingerir` exigia as faixas (ini, fim, título) explícitas.
Agora, sem faixas informadas, `detectar_faixas` varre a aba e acha os
cabeçalhos por heurística: candidata é uma linha com rótulo de texto e SEM
valor numérico (uma linha de item sempre tem rótulo+valor). Só é aceita como
seção real se render itens E uma linha de total para ancorar — mesma
filosofia de "ancorar no total" já usada na reconciliação.

Usa uma planilha sintética (criada e descartada no teste) em vez do HOPE
real, para não depender de um arquivo binário no repo nem do ambiente onde
o HOPE foi carregado originalmente.
"""
import os
import tempfile

import openpyxl

from mef_engine.ingest.planilha import detectar_e_ingerir, detectar_faixas


def _planilha_sintetica():
    """2 seções de CAPEX (com cabeçalho, itens e total) + 1 linha de ruído
    (rótulo de texto solto, sem total nem itens depois) + 1 linha de
    parâmetro (cai no filtro de PALAVRAS_RUIDO)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    linhas = [
        ("Taxa de desconto anual",),       # ruído (PALAVRAS_RUIDO: "taxa")
        ("CAPEX Complexo Hospitalar",),    # cabeçalho 1
        ("Obras civis", 1000.0),
        ("Equipamentos", 500.0),
        ("Total CAPEX Complexo Hospitalar", 1500.0),
        ("Nota: valores em R$ mil",),      # ruído: texto solto, sem total depois
        ("CAPEX LACEN",),                  # cabeçalho 2
        ("Reforma", 200.0),
        ("Mobiliário", 100.0),
        ("Total CAPEX LACEN", 300.0),
    ]
    for i, linha in enumerate(linhas, start=1):
        ws.cell(row=i, column=1, value=linha[0])
        if len(linha) > 1:
            ws.cell(row=i, column=2, value=linha[1])
    return wb


def test_detectar_faixas_acha_as_duas_secoes_e_ignora_ruido():
    wb = _planilha_sintetica()
    ws = wb.active
    faixas = detectar_faixas(ws)
    titulos = [f[2] for f in faixas]
    assert titulos == ["CAPEX Complexo Hospitalar", "CAPEX LACEN"]
    assert "Nota: valores em R$ mil" not in titulos
    print(f"  [1] Detectou as 2 seções reais e descartou o ruído: {titulos}  OK")


def test_faixas_detectadas_reconciliam():
    wb = _planilha_sintetica()
    ws = wb.active
    faixas = detectar_faixas(ws)
    ini, fim, titulo = faixas[0]
    assert (ini, fim, titulo) == (3, 5, "CAPEX Complexo Hospitalar")
    print(f"  [2] Faixa 1 detectada: linhas {ini}-{fim} ('{titulo}')  OK")


def test_detectar_e_ingerir_sem_faixas_explicitas():
    wb = _planilha_sintetica()
    with tempfile.TemporaryDirectory() as tmp:
        caminho = os.path.join(tmp, "sintetica.xlsx")
        wb.save(caminho)
        secs = detectar_e_ingerir(caminho, wb.active.title)  # faixas=None -> auto
    assert len(secs) == 2
    rec0 = secs[0].reconciliar()
    rec1 = secs[1].reconciliar()
    assert rec0["ok"] and rec1["ok"]
    assert secs[0].soma_itens == 1500.0
    assert secs[1].soma_itens == 300.0
    print(f"  [3] detectar_e_ingerir sem faixas explícitas: 2 seções, ambas reconciliam  OK")


def test_faixas_explicitas_continuam_funcionando():
    wb = _planilha_sintetica()
    with tempfile.TemporaryDirectory() as tmp:
        caminho = os.path.join(tmp, "sintetica.xlsx")
        wb.save(caminho)
        secs = detectar_e_ingerir(caminho, wb.active.title,
                                  faixas=[(3, 5, "CAPEX Complexo Hospitalar")])
    assert len(secs) == 1
    assert secs[0].soma_itens == 1500.0
    print("  [4] Faixas explícitas continuam tendo precedência sobre a detecção  OK")


if __name__ == "__main__":
    print("Validação da detecção automática de faixas por cabeçalho\n" + "-" * 48)
    test_detectar_faixas_acha_as_duas_secoes_e_ignora_ruido()
    test_faixas_detectadas_reconciliam()
    test_detectar_e_ingerir_sem_faixas_explicitas()
    test_faixas_explicitas_continuam_funcionando()
    print("-" * 48 + "\nTodos os testes passaram.")
