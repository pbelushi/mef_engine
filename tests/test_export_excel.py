"""
Validação da exportação Excel (v4): estrutura em alto nível (Capa /
Premissas / Projeções / Ativo Financeiro / Financiamento / Resultados /
Painel de Controle), sem referência a nenhum projeto/template de terceiros.

Diferença central para a v3: todo valor DERIVADO (impostos, FCFF/FCFE,
rolagem de dívida, ativo financeiro, TIR/VPL) entra como FÓRMULA do Excel —
só "Projeções" (CAPEX/OPEX/receita já distribuídos por período) entra como
valor. As provas-chave:
  - as abas aparecem/somem conforme regime contábil e existência de dívida;
  - as células de "Projeções" trazem o número (não fórmula);
  - as células de "Resultados"/"Painel de Controle" trazem FÓRMULA (não o
    número já calculado pelo motor);
  - se o pacote opcional `formulas` estiver disponível, RECALCULA as
    fórmulas e confere que batem com o ResultadoMEF — a prova de que a
    fórmula está certa, não só presente.
  - exportar_excel salva um .xlsx válido (round-trip por disco).
"""
import os
import tempfile
from datetime import date

from mef_engine.core import Periodo, RegimeContabil, TipoConcessao
from mef_engine.engine import calcular
from mef_engine.export import exportar_excel, montar_workbook
from mef_engine.schema import (
    Aporte, BlocoSetorial, EstruturaCapital, InputMEF, LinhaCAPEX, LinhaOPEX,
    LinhaReceitaFixa, LinhaReceitaVolume, RegimeLucro, Timing, Tributos,
)


def _bifurcado_com_divida():
    n = 6
    bloco = BlocoSetorial(
        capex=[LinhaCAPEX("Obra", 200.0)], opex=[LinhaOPEX("Opex", 5.0)],
        receitas_fixas=[LinhaReceitaFixa("CP", 30.0)],
        receitas_volume=[LinhaReceitaVolume("Tarifa", tarifa=2.0, volume=[40.0] * 4)],
        aporte=Aporte(valor_total=50.0),
    )
    inp = InputMEF(
        projeto="teste-export", periodo=Periodo.anual,
        tipo_concessao=TipoConcessao.patrocinada, regime_contabil=RegimeContabil.bifurcado,
        timing=Timing(date(2024, 1, 1), date(2024, 1, 1), n, date(2026, 1, 1)),
        capital=EstruturaCapital(taxa_desconto_anual=0.08, equity_pct_capex=0.5,
                                 taxa_juros_divida_anual=0.05),
        bloco=bloco, tributos=Tributos(),
    )
    return inp, calcular(inp)


def _comum_intangivel_sem_divida():
    n = 4
    bloco = BlocoSetorial(
        capex=[LinhaCAPEX("Obra", 100.0)], opex=[LinhaOPEX("Opex", 0.0)],
        receitas_volume=[LinhaReceitaVolume("Tarifa", tarifa=1.0, volume=[50.0] * n)],
    )
    inp = InputMEF(
        projeto="teste-export-simples", periodo=Periodo.anual,
        tipo_concessao=TipoConcessao.comum, regime_contabil=RegimeContabil.intangivel,
        timing=Timing(date(2024, 1, 1), date(2024, 1, 1), n, date(2024, 1, 1)),
        capital=EstruturaCapital(taxa_desconto_anual=0.08, equity_pct_capex=1.0),
        bloco=bloco, tributos=Tributos(),
    )
    return inp, calcular(inp)


def test_todas_as_abas_quando_ha_af_e_divida():
    inp, res = _bifurcado_com_divida()
    wb = montar_workbook(inp, res)
    assert wb.sheetnames == ["Capa", "Painel de Controle", "Premissas", "Projeções",
                             "Financiamento", "Ativo Financeiro", "Resultados"]
    print(f"  [1] Abas com AF e dívida: {wb.sheetnames}  OK")


def test_abas_financiamento_e_af_omitidas_sem_divida_e_intangivel():
    inp, res = _comum_intangivel_sem_divida()
    wb = montar_workbook(inp, res)
    assert "Financiamento" not in wb.sheetnames
    assert "Ativo Financeiro" not in wb.sheetnames
    assert wb.sheetnames == ["Capa", "Painel de Controle", "Premissas", "Projeções", "Resultados"]
    print(f"  [2] Sem AF/dívida: abas = {wb.sheetnames}  OK")


def test_projecoes_traz_valor_resultados_traz_formula():
    inp, res = _bifurcado_com_divida()
    wb = montar_workbook(inp, res)

    ws_proj = wb["Projeções"]
    cabecalhos = [c.value for c in ws_proj[1]]
    col_capex = cabecalhos.index("CAPEX") + 1
    assert isinstance(ws_proj.cell(row=2, column=col_capex).value, (int, float))
    assert abs(ws_proj.cell(row=2, column=col_capex).value - float(res.capex[0])) < 1e-9

    ws_res = wb["Resultados"]
    cabecalhos_res = [c.value for c in ws_res[1]]
    col_fcff = cabecalhos_res.index("FCFF") + 1
    col_fcfe = cabecalhos_res.index("FCFE") + 1
    valor_fcff = ws_res.cell(row=2, column=col_fcff).value
    valor_fcfe = ws_res.cell(row=2, column=col_fcfe).value
    assert isinstance(valor_fcff, str) and valor_fcff.startswith("=")
    assert isinstance(valor_fcfe, str) and valor_fcfe.startswith("=")

    ws_painel = wb["Painel de Controle"]
    rotulos = {ws_painel.cell(row=r, column=1).value: r for r in range(1, ws_painel.max_row + 1)}
    tir_cel = ws_painel.cell(row=rotulos["TIR-FCFF (período)"], column=2).value
    assert isinstance(tir_cel, str) and tir_cel.startswith("=IRR(")
    print("  [3] 'Projeções' traz valor, 'Resultados'/'Painel de Controle' trazem fórmula  OK")


def _checar_com_recalculo(nome_arquivo, inp, res, wb):
    """Se o pacote opcional `formulas` estiver instalado, recalcula o .xlsx
    gerado e confere contra o ResultadoMEF — sem ele, só valida estrutura
    (a prova de que as fórmulas estão certas, não só presentes)."""
    try:
        import formulas
    except ImportError:
        print("  (pacote opcional 'formulas' ausente — recálculo pulado)")
        return
    with tempfile.TemporaryDirectory() as tmp:
        caminho = os.path.join(tmp, nome_arquivo)
        wb.save(caminho)
        xl = formulas.ExcelModel().loads(caminho).finish()
        sol = xl.calculate()

        def get(aba, celula):
            chave = f"'[{nome_arquivo}]{aba.upper()}'!{celula}"
            return float(sol[chave].value[0, 0])

        n = res.malha.n_periodos
        for t in range(n):
            xlv = get("Resultados", f"M{2+t}")
            assert abs(xlv - res.fcff[t]) < 1e-6, f"FCFF[{t}]: motor={res.fcff[t]} excel={xlv}"
            xlv = get("Resultados", f"N{2+t}")
            assert abs(xlv - res.fcfe[t]) < 1e-6, f"FCFE[{t}]: motor={res.fcfe[t]} excel={xlv}"

        ws_painel = wb["Painel de Controle"]
        rotulos = {ws_painel.cell(row=r, column=1).value: r for r in range(1, ws_painel.max_row + 1)}
        if "TIR-FCFF (período)" in rotulos:
            xlv = get("Painel de Controle", f"B{rotulos['TIR-FCFF (período)']}")
            assert abs(xlv - res.tir_fcff_periodo) < 1e-6
        xlv = get("Painel de Controle", f"B{rotulos['VPL-FCFF']}")
        assert abs(xlv - res.vpl_fcff) < 1e-6
        if res.taxa_ativo is not None and "Taxa Ativo IFRIC (por período)" in rotulos:
            xlv = get("Painel de Controle", f"B{rotulos['Taxa Ativo IFRIC (por período)']}")
            assert abs(xlv - res.taxa_ativo) < 1e-6


def test_recalculo_bate_com_resultado_bifurcado_com_divida():
    inp, res = _bifurcado_com_divida()
    wb = montar_workbook(inp, res)
    _checar_com_recalculo("bifurcado.xlsx", inp, res, wb)
    print("  [4] Recálculo das fórmulas bate com o ResultadoMEF (bifurcado + dívida)  OK")


def test_recalculo_bate_com_resultado_intangivel_sem_divida():
    inp, res = _comum_intangivel_sem_divida()
    wb = montar_workbook(inp, res)
    _checar_com_recalculo("intangivel.xlsx", inp, res, wb)
    print("  [5] Recálculo das fórmulas bate com o ResultadoMEF (intangível, sem dívida)  OK")


def test_recalculo_bate_com_resultado_presumido():
    n = 6
    bloco = BlocoSetorial(
        capex=[LinhaCAPEX("Obra", 200.0)], opex=[LinhaOPEX("Opex", 5.0)],
        receitas_fixas=[LinhaReceitaFixa("CP", 30.0)],
        receitas_volume=[LinhaReceitaVolume("Tarifa", tarifa=2.0, volume=[40.0] * 4)],
    )
    inp = InputMEF(
        projeto="presumido", periodo=Periodo.anual,
        tipo_concessao=TipoConcessao.patrocinada, regime_contabil=RegimeContabil.bifurcado,
        timing=Timing(date(2024, 1, 1), date(2024, 1, 1), n, date(2026, 1, 1)),
        capital=EstruturaCapital(taxa_desconto_anual=0.08, equity_pct_capex=1.0),
        bloco=bloco, tributos=Tributos(regime_lucro=RegimeLucro.presumido),
    )
    res = calcular(inp)
    wb = montar_workbook(inp, res)
    _checar_com_recalculo("presumido.xlsx", inp, res, wb)
    print("  [6] Recálculo das fórmulas bate com o ResultadoMEF (lucro presumido)  OK")


def test_exportar_excel_salva_arquivo_valido():
    inp, res = _bifurcado_com_divida()
    with tempfile.TemporaryDirectory() as tmp:
        caminho = os.path.join(tmp, "mef.xlsx")
        exportar_excel(inp, res, caminho)
        assert os.path.exists(caminho)
        import openpyxl
        wb2 = openpyxl.load_workbook(caminho)
        assert wb2.sheetnames == ["Capa", "Painel de Controle", "Premissas", "Projeções",
                                  "Financiamento", "Ativo Financeiro", "Resultados"]
        assert wb2["Capa"].cell(row=2, column=2).value == "teste-export"
    print("  [7] exportar_excel salva um .xlsx válido e relido confere  OK")


if __name__ == "__main__":
    print("Validação da exportação Excel\n" + "-" * 48)
    test_todas_as_abas_quando_ha_af_e_divida()
    test_abas_financiamento_e_af_omitidas_sem_divida_e_intangivel()
    test_projecoes_traz_valor_resultados_traz_formula()
    test_recalculo_bate_com_resultado_bifurcado_com_divida()
    test_recalculo_bate_com_resultado_intangivel_sem_divida()
    test_recalculo_bate_com_resultado_presumido()
    test_exportar_excel_salva_arquivo_valido()
    print("-" * 48 + "\nTodos os testes passaram.")
