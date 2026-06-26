"""
Exportação dos resultados do MEF para Excel.

Organização em alto nível espelha a de um MEF profissional comum (Capa /
Premissas / Projeções / Ativo Financeiro / Financiamento / Resultados /
Painel de Controle) — sem nenhuma referência a projeto, cor ou layout de
terceiros: é uma estrutura genérica, não uma cópia de um modelo específico.

Diferença central para a v1 (que só despejava os números já calculados pelo
motor): todo valor DERIVADO (impostos, rolagem de dívida, ativo financeiro,
FCFF/FCFE, TIR/VPL) entra como FÓRMULA do Excel, referenciando as células de
premissas/projeções — abrindo o .xlsx, dá para auditar e até alterar uma
premissa e ver o resultado recalcular, como em qualquer MEF profissional.
Só a aba "Projeções" (CAPEX/OPEX/receita já distribuídos por período a
partir das linhas de entrada — curva de desembolso, indexação) entra como
valor: é o cronograma resolvido, equivalente a uma aba de premissas/
operacional, não ao bloco de resultado.
"""
from __future__ import annotations

import numpy as np
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from ..core import MalhaTemporal, RegimeContabil
from ..engine import ResultadoMEF
from ..modules import separar_receita_por_regime, vetor_capex_creditavel, vetor_opex_creditavel
from ..schema import InputMEF, RegimeLucro

FONTE_CABECALHO = Font(bold=True)
FORMATO_NUMERO = "#,##0.00"
FORMATO_PCT = "0.0000%"
FORMATO_DATA = "dd/mm/yyyy"

PRIMEIRA_LINHA = 2  # toda aba por período: cabeçalho na linha 1, dados a partir da 2


def _nativo(v):
    """numpy scalar (np.float64 etc.) -> tipo Python nativo; openpyxl não
    lida bem com tipos numpy."""
    return v.item() if isinstance(v, np.generic) else v


def _ultima_linha(n: int) -> int:
    return PRIMEIRA_LINHA + n - 1


class _Premissas:
    """Aba de premissas: tabela rótulo/valor, uma premissa por linha. Guarda
    o endereço de cada célula escrita (`ref[nome]`) para que fórmulas em
    outras abas referenciem por nome em vez de número de linha fixo —
    evita que inserir uma premissa no meio quebre referências já escritas."""

    def __init__(self, wb):
        self.ws = wb.create_sheet("Premissas")
        self.ws.column_dimensions["A"].width = 38
        self.ws.column_dimensions["B"].width = 22
        self._linha = 1
        self.ref: dict[str, str] = {}

    def add(self, nome: str, rotulo: str, valor, formato: str | None = None) -> str:
        r = self._linha
        self.ws.cell(row=r, column=1, value=rotulo).font = FONTE_CABECALHO
        cel = self.ws.cell(row=r, column=2, value=_nativo(valor))
        if formato:
            cel.number_format = formato
        self.ref[nome] = f"Premissas!$B${r}"
        self._linha += 1
        return self.ref[nome]

    def secao(self, titulo: str):
        self.ws.cell(row=self._linha, column=1, value=titulo).font = Font(bold=True, italic=True)
        self._linha += 1


def _escrever_cabecalho_aba(ws, n: int, malha: MalhaTemporal,
                            colunas: list[tuple[str, list, str | None]]) -> dict[str, str]:
    """Escreve a coluna A (Período, 0-based) e B (Data) e, a seguir, as
    colunas declaradas em `colunas` = [(rotulo, valores_ou_formulas, formato), ...].
    `valores_ou_formulas[t]` é um valor literal OU uma string iniciada por
    '=' (fórmula). Devolve {rotulo_normalizado: letra_da_coluna} para uso em
    fórmulas de outras abas."""
    ws.cell(row=1, column=1, value="Período").font = FONTE_CABECALHO
    ws.cell(row=1, column=2, value="Data").font = FONTE_CABECALHO
    for t in range(n):
        r = PRIMEIRA_LINHA + t
        ws.cell(row=r, column=1, value=t)
        cel_data = ws.cell(row=r, column=2, value=malha.datas_inicio[t])
        cel_data.number_format = FORMATO_DATA
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 14

    letras: dict[str, str] = {}
    for c, (rotulo, valores, formato) in enumerate(colunas, start=3):
        letra = get_column_letter(c)
        letras[rotulo] = letra
        ws.cell(row=1, column=c, value=rotulo).font = FONTE_CABECALHO
        for t in range(n):
            r = PRIMEIRA_LINHA + t
            valor = valores[t]
            cel = ws.cell(row=r, column=c,
                          value=valor if isinstance(valor, str) else _nativo(valor))
            if formato:
                cel.number_format = formato
        ws.column_dimensions[letra].width = 18
    return letras


# --- Capa --------------------------------------------------------------
def _aba_capa(wb, inp: InputMEF):
    ws = wb.create_sheet("Capa")
    ws.cell(row=2, column=2, value=inp.projeto).font = Font(bold=True, size=16)
    ws.cell(row=3, column=2, value="Modelo Econômico-Financeiro").font = Font(size=12)
    ws.column_dimensions["B"].width = 50
    return ws


# --- Premissas -----------------------------------------------------------
def _aba_premissas(wb, inp: InputMEF, malha: MalhaTemporal) -> _Premissas:
    p = _Premissas(wb)
    trib = inp.tributos
    cap = inp.capital
    idx_op = malha.indice_da_data(inp.timing.inicio_operacao)
    n = malha.n_periodos

    p.secao("Projeto")
    p.add("projeto", "Projeto", inp.projeto)
    p.add("tipo_concessao", "Tipo de concessão", inp.tipo_concessao.value)
    p.add("regime_contabil", "Regime contábil", inp.regime_contabil.value)
    p.add("periodicidade", "Periodicidade", inp.periodo.value)

    p.secao("Cronograma")
    p.add("data_base", "Data-base", inp.timing.data_base, FORMATO_DATA)
    p.add("inicio_ppp", "Início da PPP/concessão", inp.timing.inicio_ppp, FORMATO_DATA)
    p.add("n_periodos", "Prazo (períodos)", n)
    p.add("inicio_operacao", "Início da operação", inp.timing.inicio_operacao, FORMATO_DATA)
    p.add("idx_op", "Índice do início da operação (0-based)", idx_op)
    p.add("por_ano", "Períodos por ano", inp.periodo.por_ano)

    p.secao("Taxa de desconto")
    p.add("taxa_desconto", "Taxa de desconto (por período)", inp.taxa_desconto_periodo, FORMATO_PCT)

    p.secao("Tributos")
    p.add("regime_lucro", "Regime de lucro", trib.regime_lucro.value)
    p.add("aplica_ircsll", "Aplica IR/CSLL (1/0)", 1 if trib.aplica_ir_csll else 0)
    p.add("aliquota_indireta", "Alíquota indireta (PIS/COFINS/ISS ou CBS/IBS)",
         trib.aliquota_indireta, FORMATO_PCT)
    p.add("aliquota_credito", "Alíquota de crédito sobre insumos",
         trib.aliquota_credito_insumos, FORMATO_PCT)
    p.add("credito_pis_cofins", "Aplica crédito sobre insumos (1/0)",
         1 if trib.credito_pis_cofins else 0)
    p.add("irpj", "IRPJ", trib.irpj, FORMATO_PCT)
    p.add("irpj_adicional", "IRPJ adicional", trib.irpj_adicional, FORMATO_PCT)
    p.add("csll", "CSLL", trib.csll, FORMATO_PCT)
    p.add("compensacao_prejuizo", "Compensa prejuízo fiscal (1/0)",
         1 if (trib.regime_lucro is RegimeLucro.real and trib.compensacao_prejuizo) else 0)
    p.add("trava_compensacao", "Trava de compensação de prejuízo", trib.trava_compensacao_prejuizo, FORMATO_PCT)
    p.add("aporte_tributavel", "Aporte compõe a base tributável (1/0)",
         1 if trib.aporte_tributavel else 0)
    if trib.regime_lucro is RegimeLucro.presumido:
        from ..schema import preset_por_atividade
        preset_at = preset_por_atividade(trib.atividade_economica)
        p.add("presuncao_irpj", "Presunção de lucro (IRPJ)", preset_at["presuncao_irpj"], FORMATO_PCT)
        p.add("presuncao_csll", "Presunção de lucro (CSLL)", preset_at["presuncao_csll"], FORMATO_PCT)

    p.secao("Financiamento")
    p.add("equity_pct", "Equity % do CAPEX", cap.equity_pct_capex, FORMATO_PCT)
    taxa_divida_periodo = (1 + cap.taxa_juros_divida_anual) ** (1 / inp.periodo.por_ano) - 1
    p.add("taxa_divida", "Taxa de juros da dívida (por período)", taxa_divida_periodo, FORMATO_PCT)
    tem_divida = cap.equity_pct_capex < 1.0 and inp.capex_total > 0
    if tem_divida:
        prazo_amort = cap.prazo_amortizacao_periodos or (n - idx_op)
        prazo_amort = max(1, min(prazo_amort, n - idx_op))
        p.add("prazo_amort", "Prazo de amortização (períodos)", prazo_amort)

    p.secao("Ativo financeiro (IFRIC 12)")
    if inp.regime_contabil is RegimeContabil.bifurcado:
        if inp.fracao_ativo_financeiro is not None:
            p.add("fracao_af", "Fração de CAPEX/OPEX no ativo financeiro",
                 inp.fracao_ativo_financeiro, FORMATO_PCT)
        # quando derivada (None), a célula é uma FÓRMULA escrita depois de
        # montar a aba Projeções (depende da soma de receita garantida/risco) —
        # ver `_aba_projecoes`/`_completar_fracao_af_derivada`.

    return p


# --- Projeções (cronograma já resolvido: CAPEX/OPEX/receita por período) ---
def _aba_projecoes(wb, inp: InputMEF, malha: MalhaTemporal, resultado: ResultadoMEF, premissas: _Premissas):
    n = malha.n_periodos
    ws = wb.create_sheet("Projeções")
    rec = separar_receita_por_regime(inp, malha)
    capex_cred = vetor_capex_creditavel(inp, malha)
    opex_cred = vetor_opex_creditavel(inp, malha)

    # ordem fixa das colunas (a partir da coluna 3 = C); calculada antes de
    # escrever, para que "Receita Total" referencie as letras certas das
    # colunas de Receita Garantida/Risco que vêm antes dela na mesma lista.
    nomes = ["CAPEX", "OPEX", "Receita Garantida", "Receita Risco de Demanda",
            "Receita Total", "Aporte", "CAPEX Creditável", "OPEX Creditável"]
    col = {nome: get_column_letter(c) for c, nome in enumerate(nomes, start=3)}

    letras = _escrever_cabecalho_aba(ws, n, malha, [
        ("CAPEX", list(resultado.capex), FORMATO_NUMERO),
        ("OPEX", list(resultado.opex), FORMATO_NUMERO),
        ("Receita Garantida", list(rec["garantida"]), FORMATO_NUMERO),
        ("Receita Risco de Demanda", list(rec["risco_demanda"]), FORMATO_NUMERO),
        ("Receita Total",
         [f"={col['Receita Garantida']}{PRIMEIRA_LINHA+t}+{col['Receita Risco de Demanda']}{PRIMEIRA_LINHA+t}"
          for t in range(n)], FORMATO_NUMERO),
        ("Aporte", list(resultado.aporte), FORMATO_NUMERO),
        ("CAPEX Creditável", list(capex_cred), FORMATO_NUMERO),
        ("OPEX Creditável", list(opex_cred), FORMATO_NUMERO),
    ])
    assert letras == col, f"letras calculadas não batem com as escritas: {letras} != {col}"
    return ws, letras


def _completar_fracao_af_derivada(premissas: _Premissas, n: int, letras_proj: dict):
    """Para o regime bifurcado sem override explícito, a fração de ativo
    financeiro é DERIVADA da mistura de receita (garantida / total) — escreve
    essa derivação como fórmula na aba Premissas, em vez de número fixo."""
    ultima = _ultima_linha(n)
    g = letras_proj["Receita Garantida"]
    r = letras_proj["Receita Risco de Demanda"]
    formula = (f"=SUM(Projeções!{g}{PRIMEIRA_LINHA}:{g}{ultima})/"
              f"(SUM(Projeções!{g}{PRIMEIRA_LINHA}:{g}{ultima})+"
              f"SUM(Projeções!{r}{PRIMEIRA_LINHA}:{r}{ultima}))")
    rotulo = "Fração de CAPEX/OPEX no ativo financeiro (derivada)"
    return premissas.add("fracao_af", rotulo, formula, FORMATO_PCT)


# --- Financiamento ---------------------------------------------------------
def _aba_financiamento(wb, inp: InputMEF, malha: MalhaTemporal, premissas: _Premissas, letras_proj: dict):
    cap = inp.capital
    if not (cap.equity_pct_capex < 1.0 and inp.capex_total > 0):
        return None, None
    n = malha.n_periodos
    ws = wb.create_sheet("Financiamento")
    ultima = _ultima_linha(n)
    capex_col = letras_proj["CAPEX"]
    idx_op = malha.indice_da_data(inp.timing.inicio_operacao)
    idx_op_ref = premissas.ref["idx_op"]
    taxa_ref = premissas.ref["taxa_divida"]
    equity_ref = premissas.ref["equity_pct"]
    prazo_ref = premissas.ref["prazo_amort"]

    # colunas: C=Saque D=Saldo Inicial E=Juros F=Amortização G=Serviço da Dívida H=Saldo Final
    col_saldo_ini, col_juros, col_amort, col_saldo_fim = "D", "E", "F", "H"
    saque = [f"=Projeções!{capex_col}{PRIMEIRA_LINHA+t}*(1-{equity_ref})" for t in range(n)]
    saldo_ini, juros, amort, servico, saldo_fim = [], [], [], [], []
    for t in range(n):
        r = PRIMEIRA_LINHA + t
        # saldo inicial do período = saldo final do período anterior (0 na 1ª linha)
        si = f"=IF(ROW()={PRIMEIRA_LINHA},0,{col_saldo_fim}{r-1})"
        j = (f"=IF(A{r}<{idx_op_ref},"
            f"({taxa_ref}/2*(2*{col_saldo_ini}{r}+C{r}))/(1-{taxa_ref}/2),"
            f"{col_saldo_ini}{r}*{taxa_ref})")
        # amortização referencia a célula auxiliar $K$1 (saldo no início da
        # operação / prazo) — e $K$1, por sua vez, referencia o Saldo Inicial
        # exatamente na linha de início da operação. Se TODA linha (inclusive
        # as de construção) tivesse uma fórmula textualmente mencionando
        # $K$1, a cadeia Saldo Inicial->Saldo Final->Amortização->$K$1
        # fecharia um ciclo (mesmo sendo 0 em valor nas linhas de construção,
        # a referência textual já basta para o Excel marcar como circular).
        # Por isso linhas de construção (t<idx_op, decidido aqui em tempo de
        # geração, não dinamicamente na planilha) entram com 0 literal.
        if t < idx_op:
            a = 0
        else:
            a = f"=IF(A{r}<({idx_op_ref}+{prazo_ref}),MIN($K$1,{col_saldo_ini}{r}),0)"
        sd = f"=IF(A{r}>={idx_op_ref},{col_juros}{r}+{col_amort}{r},0)"
        sf = f"={col_saldo_ini}{r}+C{r}+IF(A{r}<{idx_op_ref},{col_juros}{r},0)-{col_amort}{r}"
        saldo_ini.append(si); juros.append(j); amort.append(a); servico.append(sd); saldo_fim.append(sf)

    letras = _escrever_cabecalho_aba(ws, n, malha, [
        ("Saque", saque, FORMATO_NUMERO),
        ("Saldo Inicial", saldo_ini, FORMATO_NUMERO),
        ("Juros", juros, FORMATO_NUMERO),
        ("Amortização", amort, FORMATO_NUMERO),
        ("Serviço da Dívida", servico, FORMATO_NUMERO),
        ("Saldo Final", saldo_fim, FORMATO_NUMERO),
    ])
    # auxiliar: amortização constante do período = saldo no início da operação /
    # prazo. Referencia a célula ESPECÍFICA da linha de início da operação (não
    # um INDEX sobre a faixa inteira D2:D{ultima}) porque um INDEX sobre a
    # faixa criaria dependência circular: linhas posteriores da própria coluna
    # Saldo Inicial dependem do Saldo Final, que depende da Amortização, que
    # depende desta célula auxiliar — um INDEX sobre toda a faixa faria esta
    # célula "depender de si mesma" via essas linhas, mesmo sem precisar do
    # valor delas. `idx_op` já é conhecido em tempo de geração (não muda
    # dinamicamente na planilha), então a linha exata pode ser resolvida aqui.
    linha_idx_op = PRIMEIRA_LINHA + idx_op
    ws.cell(row=1, column=10, value="Amortização por período (auxiliar)").font = FONTE_CABECALHO
    ws.cell(row=1, column=11,
           value=f"={col_saldo_ini}{linha_idx_op}/{prazo_ref}"
           ).number_format = FORMATO_NUMERO
    ws.column_dimensions["J"].width = 32
    assert letras == {"Saque": "C", "Saldo Inicial": "D", "Juros": "E", "Amortização": "F",
                      "Serviço da Dívida": "G", "Saldo Final": "H"}
    return ws, letras


# --- Ativo Financeiro (IFRIC 12) -------------------------------------------
def _aba_ativo_financeiro(wb, inp: InputMEF, malha: MalhaTemporal, premissas: _Premissas, letras_proj: dict):
    if inp.regime_contabil is RegimeContabil.intangivel:
        return None, None
    n = malha.n_periodos
    ws = wb.create_sheet("Ativo Financeiro")
    ultima = _ultima_linha(n)
    capex_col, opex_col = letras_proj["CAPEX"], letras_proj["OPEX"]

    if inp.regime_contabil is RegimeContabil.bifurcado:
        fracao_ref = premissas.ref["fracao_af"]
        capex_af = [f"=Projeções!{capex_col}{PRIMEIRA_LINHA+t}*{fracao_ref}" for t in range(n)]
        opex_af = [f"=Projeções!{opex_col}{PRIMEIRA_LINHA+t}*{fracao_ref}" for t in range(n)]
        contrap_col = letras_proj["Receita Garantida"]
        contrap_af = [f"=Projeções!{contrap_col}{PRIMEIRA_LINHA+t}" for t in range(n)]
    else:  # ativo_financeiro puro: 100% do CAPEX/OPEX e da receita total
        capex_af = [f"=Projeções!{capex_col}{PRIMEIRA_LINHA+t}" for t in range(n)]
        opex_af = [f"=Projeções!{opex_col}{PRIMEIRA_LINHA+t}" for t in range(n)]
        contrap_col = letras_proj["Receita Total"]
        contrap_af = [f"=Projeções!{contrap_col}{PRIMEIRA_LINHA+t}" for t in range(n)]

    # colunas: C=CAPEX AF D=OPEX AF E=Contraprestação AF F=Fluxo Aux G=AF Inicial H=Receita Financeira I=AF Final
    fluxo_aux = [f"=E{PRIMEIRA_LINHA+t}-C{PRIMEIRA_LINHA+t}-D{PRIMEIRA_LINHA+t}" for t in range(n)]
    af_ini, recfin, af_fim = [], [], []
    for t in range(n):
        r = PRIMEIRA_LINHA + t
        # AF inicial do período = AF final do período anterior (0 na 1ª linha)
        af_ini.append(f"=IF(ROW()={PRIMEIRA_LINHA},0,I{r-1})")
        recfin.append(f"=G{r}*$L$1")
        af_fim.append(f"=G{r}+C{r}+D{r}-E{r}+H{r}")

    letras = _escrever_cabecalho_aba(ws, n, malha, [
        ("CAPEX AF", capex_af, FORMATO_NUMERO),
        ("OPEX AF", opex_af, FORMATO_NUMERO),
        ("Contraprestação AF", contrap_af, FORMATO_NUMERO),
        ("Fluxo Auxiliar (p/ taxa implícita)", fluxo_aux, FORMATO_NUMERO),
        ("AF Inicial", af_ini, FORMATO_NUMERO),
        ("Receita Financeira", recfin, FORMATO_NUMERO),
        ("AF Final", af_fim, FORMATO_NUMERO),
    ])
    ws.cell(row=1, column=11, value="Taxa Ativo IFRIC (por período, auxiliar)").font = FONTE_CABECALHO
    ws.cell(row=1, column=12,
           value=f"=IRR(F{PRIMEIRA_LINHA}:F{ultima},0.05)").number_format = FORMATO_PCT
    ws.column_dimensions["K"].width = 32
    assert letras == {"CAPEX AF": "C", "OPEX AF": "D", "Contraprestação AF": "E",
                      "Fluxo Auxiliar (p/ taxa implícita)": "F", "AF Inicial": "G",
                      "Receita Financeira": "H", "AF Final": "I"}
    return ws, letras


# --- Resultados (impostos, FCFF, FCFE) -------------------------------------
def _aba_resultados(wb, inp: InputMEF, malha: MalhaTemporal, premissas: _Premissas,
                    letras_proj: dict, letras_fin: dict | None):
    n = malha.n_periodos
    trib = inp.tributos
    ws = wb.create_sheet("Resultados")
    ultima = _ultima_linha(n)

    capex_col, opex_col = letras_proj["CAPEX"], letras_proj["OPEX"]
    receita_col, aporte_col = letras_proj["Receita Total"], letras_proj["Aporte"]
    capex_cred_col, opex_cred_col = letras_proj["CAPEX Creditável"], letras_proj["OPEX Creditável"]
    aporte_trib_ref = premissas.ref["aporte_tributavel"]
    credito_ref = premissas.ref["credito_pis_cofins"]
    aliq_indireta_ref = premissas.ref["aliquota_indireta"]
    aliq_credito_ref = premissas.ref["aliquota_credito"]
    aplica_ircsll_ref = premissas.ref["aplica_ircsll"]
    irpj_ref, irpj_ad_ref, csll_ref = premissas.ref["irpj"], premissas.ref["irpj_adicional"], premissas.ref["csll"]

    receita_trib, creditos, indiretos, deprec, lucro = [], [], [], [], []
    base_irpj, base_csll, saldo_prej, ircsll, imp_total, fcff, fcfe = [], [], [], [], [], [], []

    compensa = trib.regime_lucro is RegimeLucro.real and trib.compensacao_prejuizo
    trava_ref = premissas.ref.get("trava_compensacao")
    presumido = trib.regime_lucro is RegimeLucro.presumido
    if presumido:
        presuncao_irpj_ref = premissas.ref["presuncao_irpj"]
        presuncao_csll_ref = premissas.ref["presuncao_csll"]

    for t in range(n):
        r = PRIMEIRA_LINHA + t
        receita_trib.append(f"=Projeções!{receita_col}{r}+Projeções!{aporte_col}{r}*{aporte_trib_ref}")
        creditos.append(
            f"=IF({credito_ref}=1,(Projeções!{capex_cred_col}{r}+Projeções!{opex_cred_col}{r})*{aliq_credito_ref},0)")
        indiretos.append(f"=MAX(C{r}*{aliq_indireta_ref}-D{r},0)")
        deprec.append(f"=SUM(Projeções!{capex_col}${PRIMEIRA_LINHA}:{capex_col}${ultima})/{premissas.ref['n_periodos']}")
        lucro.append(f"=C{r}-Projeções!{opex_col}{r}-F{r}-E{r}")

        if presumido:
            base_irpj.append(f"=C{r}*{presuncao_irpj_ref}")
            base_csll.append(f"=C{r}*{presuncao_csll_ref}")
            saldo_prej.append(0)
        elif compensa:
            anterior = f"IF(ROW()={PRIMEIRA_LINHA},0,J{r-1})"
            compensavel = f"MIN({anterior},G{r}*{trava_ref})"
            base_irpj.append(f"=IF(G{r}<=0,0,G{r}-{compensavel})")
            base_csll.append(f"=H{r}")
            saldo_prej.append(f"=IF(G{r}<=0,{anterior}-G{r},{anterior}-{compensavel})")
        else:
            base_irpj.append(f"=MAX(G{r},0)")
            base_csll.append(f"=H{r}")
            saldo_prej.append(0)

        ircsll.append(f"=IF({aplica_ircsll_ref}=1,H{r}*({irpj_ref}+{irpj_ad_ref})+I{r}*{csll_ref},0)")
        imp_total.append(f"=E{r}+K{r}")
        fcff.append(f"=Projeções!{receita_col}{r}-Projeções!{opex_col}{r}-Projeções!{capex_col}{r}"
                   f"+Projeções!{aporte_col}{r}-L{r}")
        if letras_fin:
            saque_col, servico_col = letras_fin["Saque"], letras_fin["Serviço da Dívida"]
            fcfe.append(f"=M{r}+Financiamento!{saque_col}{r}-Financiamento!{servico_col}{r}")
        else:
            fcfe.append(f"=M{r}")

    letras = _escrever_cabecalho_aba(ws, n, malha, [
        ("Receita Tributável", receita_trib, FORMATO_NUMERO),
        ("Créditos PIS/COFINS", creditos, FORMATO_NUMERO),
        ("Indiretos", indiretos, FORMATO_NUMERO),
        ("Depreciação", deprec, FORMATO_NUMERO),
        ("Lucro", lucro, FORMATO_NUMERO),
        ("Base IRPJ", base_irpj, FORMATO_NUMERO),
        ("Base CSLL", base_csll, FORMATO_NUMERO),
        ("Saldo Prejuízo Acumulado", saldo_prej, FORMATO_NUMERO),
        ("IR/CSLL", ircsll, FORMATO_NUMERO),
        ("Impostos Total", imp_total, FORMATO_NUMERO),
        ("FCFF", fcff, FORMATO_NUMERO),
        ("FCFE", fcfe, FORMATO_NUMERO),
    ])
    assert letras["FCFF"] == "M" and letras["FCFE"] == "N"
    return ws, letras


# --- Painel de Controle (resumo / KPIs) ------------------------------------
def _aba_painel_controle(wb, inp: InputMEF, malha: MalhaTemporal, premissas: _Premissas,
                         letras_proj: dict, letras_fin: dict | None, letras_af: dict | None,
                         letras_res: dict):
    n = malha.n_periodos
    ultima = _ultima_linha(n)
    ws = wb.create_sheet("Painel de Controle")
    taxa_ref = premissas.ref["taxa_desconto"]
    por_ano_ref = premissas.ref["por_ano"]

    def soma(aba, col):
        return f"=SUM({aba}!{col}{PRIMEIRA_LINHA}:{col}{ultima})"

    def vpl(aba, col):
        if n == 1:
            return f"={aba}!{col}{PRIMEIRA_LINHA}"
        return f"={aba}!{col}{PRIMEIRA_LINHA}+NPV({taxa_ref},{aba}!{col}{PRIMEIRA_LINHA+1}:{col}{ultima})"

    def tir(aba, col):
        if n < 2:
            return None
        return f"=IRR({aba}!{col}{PRIMEIRA_LINHA}:{col}{ultima},0.1)"

    linhas: list[tuple[str, object, str | None]] = [
        ("Projeto", f"={premissas.ref['projeto']}", None),
        ("Tipo de concessão", f"={premissas.ref['tipo_concessao']}", None),
        ("Regime contábil", f"={premissas.ref['regime_contabil']}", None),
        ("Periodicidade", f"={premissas.ref['periodicidade']}", None),
        ("Períodos", f"={premissas.ref['n_periodos']}", None),
        ("CAPEX Total", soma("Projeções", letras_proj["CAPEX"]), FORMATO_NUMERO),
        ("OPEX Total", soma("Projeções", letras_proj["OPEX"]), FORMATO_NUMERO),
        ("Receita Total", soma("Projeções", letras_proj["Receita Total"]), FORMATO_NUMERO),
        ("Aporte Total", soma("Projeções", letras_proj["Aporte"]), FORMATO_NUMERO),
    ]
    if letras_fin:
        linhas.append(("Dívida Sacada", soma("Financiamento", letras_fin["Saque"]), FORMATO_NUMERO))
    tem_tir_fcff = tir("Resultados", letras_res["FCFF"]) is not None
    if tem_tir_fcff:
        linhas.append(("TIR-FCFF (período)", tir("Resultados", letras_res["FCFF"]), FORMATO_PCT))
        linhas.append(("TIR-FCFF anual", None, FORMATO_PCT))  # resolvida abaixo (referencia a própria linha)
    linhas.append(("VPL-FCFF", vpl("Resultados", letras_res["FCFF"]), FORMATO_NUMERO))
    tem_tir_fcfe = tir("Resultados", letras_res["FCFE"]) is not None
    if tem_tir_fcfe:
        linhas.append(("TIR-FCFE (período)", tir("Resultados", letras_res["FCFE"]), FORMATO_PCT))
        linhas.append(("TIR-FCFE anual", None, FORMATO_PCT))
    linhas.append(("VPL-FCFE", vpl("Resultados", letras_res["FCFE"]), FORMATO_NUMERO))
    if letras_af:
        linhas.append(("Taxa Ativo IFRIC (por período)", "='Ativo Financeiro'!$L$1", FORMATO_PCT))
    if "fracao_af" in premissas.ref:
        linhas.append(("Fração Ativo Financeiro (efetiva)", f"={premissas.ref['fracao_af']}", FORMATO_PCT))

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22
    linha_por_chave = {}
    for r, (rotulo, valor, formato) in enumerate(linhas, start=1):
        ws.cell(row=r, column=1, value=rotulo).font = FONTE_CABECALHO
        cel = ws.cell(row=r, column=2, value=valor)
        if formato:
            cel.number_format = formato
        linha_por_chave[rotulo] = r
    # TIR anual referencia a própria linha da TIR período — só dá para
    # escrever a fórmula depois de saber em que linha ela caiu.
    if tem_tir_fcff:
        r_tir, r_anual = linha_por_chave["TIR-FCFF (período)"], linha_por_chave["TIR-FCFF anual"]
        ws.cell(row=r_anual, column=2, value=f"=(1+B{r_tir})^{por_ano_ref}-1")
    if tem_tir_fcfe:
        r_tir, r_anual = linha_por_chave["TIR-FCFE (período)"], linha_por_chave["TIR-FCFE anual"]
        ws.cell(row=r_anual, column=2, value=f"=(1+B{r_tir})^{por_ano_ref}-1")
    return ws


def montar_workbook(inp: InputMEF, resultado: ResultadoMEF) -> openpyxl.Workbook:
    """Monta o workbook em memória, sem salvar — separado de `exportar_excel`
    para testar células direto, sem round-trip por disco."""
    malha = resultado.malha
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove a aba default vazia

    _aba_capa(wb, inp)
    premissas = _aba_premissas(wb, inp, malha)
    _, letras_proj = _aba_projecoes(wb, inp, malha, resultado, premissas)
    if inp.regime_contabil is RegimeContabil.bifurcado and inp.fracao_ativo_financeiro is None:
        _completar_fracao_af_derivada(premissas, malha.n_periodos, letras_proj)
    _, letras_fin = _aba_financiamento(wb, inp, malha, premissas, letras_proj)
    _, letras_af = _aba_ativo_financeiro(wb, inp, malha, premissas, letras_proj)
    _, letras_res = _aba_resultados(wb, inp, malha, premissas, letras_proj, letras_fin)
    _aba_painel_controle(wb, inp, malha, premissas, letras_proj, letras_fin, letras_af, letras_res)

    # "Painel de Controle" é montado por último (precisa saber as referências
    # de todas as outras abas), mas pertence logo após a Capa na leitura —
    # reordena as abas sem afetar nenhuma fórmula (não depende de posição).
    ordem = ["Capa", "Painel de Controle", "Premissas", "Projeções",
            "Financiamento", "Ativo Financeiro", "Resultados"]
    wb._sheets = [wb[nome] for nome in ordem if nome in wb.sheetnames]
    return wb


def exportar_excel(inp: InputMEF, resultado: ResultadoMEF, caminho: str) -> None:
    """Exporta o MEF para um .xlsx com fórmulas vivas (Premissas, Projeções,
    Ativo Financeiro, Financiamento, Resultados, Painel de Controle) — toda
    grandeza derivada é uma fórmula auditável no próprio Excel, não um
    número já calculado pelo motor."""
    montar_workbook(inp, resultado).save(caminho)
